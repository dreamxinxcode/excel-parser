import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from datetime import datetime
import time
import os


def list_all(file):
        start_time = time.time()
        print(
            "\033[1m\033[96mOpening:\033[0m \033[1m\033[93m{0}\033[0m".format(file))
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if (cell.value != None):
                        print(cell.value)


        print(
            "\033[1m\033[96mSaving:\033[0m \033[1m\033[93m{}\033[0m in \033[1m\033[96m{} seconds\033[0m\n".format(file, round((start_time - time.time()), 2)))
        wb.save(file)


def template():
    wb = Workbook()
    ws = wb.active

    model = ws['A2']
    model.font = Font(color=colors.RED, bold=True, size=26, name='Calibri')
    model.value = "Boat Model"

    # Aluminum
    ws['B3'] = "Aluminum"
    ws['B3'].font = Font(name='Calibri', size=16, bold=True)
    header_items = ["Quantity", "Item Description", "Supplier Part #",
                    "Supplier Name", "Cost (EACH)", "Cost (TOTAL)"]
    header = ws.append(header_items)
    header.font = Font(name="Calibri", size=11, bold=True, italic=True)

    wb.save('template.xlsx')


if(__name__ == "__main__"):
    file = input("File name: ")
    list_all(file)
